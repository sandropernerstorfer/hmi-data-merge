import os
import tkinter
from tkinter.filedialog import askopenfilename
from openpyxl import Workbook, load_workbook
from database import safetyAreas
from utils import printInfoBlock


# Clear Console Window
os.system('cls||clear')
# Clear tkinter frame
root = tkinter.Tk()
root.withdraw()


# Ask for Master File-Path
filePath = askopenfilename(filetypes=[( 'Excel File', '.xlsx .xls')])
if(filePath == ''): exit()

# Hardcoded Variables for Master File - can be dynamic later
firstRow = 8
pidRow = 3

# Load Master-Workbook & Target-Sheet
print("\nLoading Excel Data from:  "+"\033[2;32;40m.../"+filePath.split('/')[-2]+'/'+filePath.split('/')[-1]+"\033[0m\n")
master_wb = load_workbook(filePath)
master_ws = master_wb.active

# User Input - Filtering Parameters
printInfoBlock('Fill out parameters to filter:')
pid = input("\nP&ID: ")
print('')
pid = int(pid)

# List for Output
entries = []

# Loop through Sheet-Rows
for key, *values in master_ws.iter_rows(min_row = firstRow):
  if(key.value==None):
    break
  if(values[pidRow-2].value == pid):
    entries.append([v.value for v in values])

# Print Result Infos
printInfoBlock('Found '+str(len(entries))+' entries.')
printInfoBlock('First: '+entries[0][4]+str(entries[0][5])+' ... Last: '+entries[-1][4]+str(entries[-1][5]))

# User Confirmation
confirmation = input('Confirmation: Process this data and create new Excel-File? [y,n]: ')
if(confirmation == 'y' or confirmation == 'yes' or confirmation == ''):
  print('Worked')
else:
  exit()

# Hardcoded Variables for Destination File - can be dynamic later
firstRow = 2

# Load Destination Workbook & Sheet
print('Creating new Excel-Workbook and importing Data ...')
wb = Workbook()
ws = wb.active
ws.title = +str(pid)+'-Data'

# rows einschreiben
for row in entries:
  ws.append(row)

# saving
wb.save(str(pid)+'-Processed.xlsx')