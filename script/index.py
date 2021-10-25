import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from assets.filtering.__filterUtils import getAllWithPid
from assets.utils import *
from assets.database import sheetName

# --------------------------------- #
def main():
  # Get desktop path and dir name
  desktopPath = os.path.expanduser("~/Desktop")
  directoryName = str(os.path.split(desktopPath)[-1])

  # Initial Console Clearing
  clearConsole()

  # Get Master File Path and load Worksheet
  filePath = getExcelPath('Master')
  master_ws = getExcelSheet(filePath, 'Master', sheetName)

  # Logic Circle Start  |  Take Input-Values, Filter Data, Save Sheet
  while True:
    
    # Input Circle Start
    while True:
      
      # PID-Input Circle Start
      pid = askForPid()

      # Get elements with given PID
      allElements = getAllWithPid(pid, master_ws)
      clearConsole()
      if(len(allElements) == 0):
        printPidFilterResult(allElements, pid, 'red')
        printListItem('No Elements with given PID found. Try searching again.', 'red')
        continue
      else: printPidFilterResult(allElements, pid, 'green')
      
      # Typical-Input Circle Start | Ask for typicals, search and get filter functions
      filterToolsList = askAndReturnFilterTools()

      # Get filtered lists with Typical-Functions and print results
      finalLists = [] # this stores all final filtered lists + typical name
      for filterTools in filterToolsList:
        filterResults = filterTools[0](allElements, filterTools[1])     # returns [entries, typical name]
        printTypicalFilterResults(filterResults[0], filterResults[1])
        if(len(filterResults) > 0):
          finalLists.append(filterResults + [filterTools[2]])

      # Check if there is at least 1 entry to continue with
      if(len(finalLists) == 0):
        printInfoBlock('Found no entries.', 'red')
        print('\nOne of following could be the reason:')
        printListItem('No entries found with given PID: '+pid, 'yellow')
        printListItem('None of the entries fall into the given typicals categories.', 'yellow')
        print('\nYou can try again using different parameters.')
      else: break
    
    # Get user confirmation for continuing with populating new sheet
    print('')
    confirmation = getUserConfirmation('Continue to \033[92mMaster | ProcessLibrary\033[0m merging stage?')
    clearConsole()
    if(confirmation == False):
      confirmation = getUserConfirmation('Start again?')
      clearConsole()
      if(confirmation == True):
        continue
      else:
        exit()

    # Instanciate Output File Object
    wb = Workbook()
    
    # File Input Circle for ProcessLib File
    while True:

      # Get Path of ProcessLib File
      filePath = getProcessPath('ProcessLibraryOnlineConfigTool')
      if(filePath == False): continue

      # Get Excel Data from ProcessLib File
      book = getProcessFileData(filePath, 'ProcessLibraryOnlineConfigTool')
      if(book == None):
        confirmation = getUserConfirmation('Want to search for the ProcessLibraryOnlineConfigTool File again ?')
        if(confirmation == True): continue
        else: exit()
      else: break

    # Loop through all saved typicals with their items
    # list = [listItems, typicalName]
    for list in finalLists:
      
      # get xls sheet by typical name
      sheet = book.sheet_by_name(list[1])
      rowCount = sheet.nrows
      # get XLS data
      xlsData = []
      for row in range(rowCount):
        xlsData.append(sheet.row_values(rowx=row))
      
      # Call Merge function for current typical and get final dataset
      doneList = list[2](list[0], xlsData)
      
      # Create Sheet on instanciated workbook object with Typical Name
      # Also create sheet with only the imported for checking
      printListItem('Creating '+list[1]+' Sheet and importing Data ...', 'cyan')
      controlSheet = wb.create_sheet(list[1]+'-Imported')
      outputSheet = wb.create_sheet(list[1]+'-Full')

      # Populate Imported checking sheet
      controlSheet.append([' '])
      controlSheet.append(['This is just the control sheet, if you want to check the master imports for '+list[1]+' again.'])
      controlSheet.merge_cells('A1:J1')
      controlSheet.append(['The '+list[1]+'-Full Sheet has to be manually imported into the ProcessLib-File '+list[1]+' Sheet.'])
      controlSheet.merge_cells('A2:J2')
      controlSheet.append([' '])
      
      for row in list[0]:
        controlSheet.append(row)
        
      # Populate final and full output sheet
      for row in doneList:
        outputSheet.append(row)

    # Save new file in 'output' folder
    try:
      defaultSheet = wb['Sheet']
      wb.remove(defaultSheet)
    except: pass

    try:
      savePath = os.path.join(desktopPath, pid+'-processed.xlsx')
      wb.save(savePath)
      print('')
      printInfoBlock('File for PID: '+pid+' saved to \''+directoryName+'\'.', 'green')
      print('')
      printListItem('Make sure you now manually merge the sheets before you start again.', 'yellow')
      printListItem('So you work with the updated ProcessedLibraryConfigTool File.', 'yellow')
      print('')
      confirmation = getUserConfirmation('Want to start again ?')
      clearConsole()
      if(confirmation == True): continue
      else: break
    except:
      clearConsole()
      print('Something went wrong while saving the file. Make sure the file you are writing to \033[93mis closed\033[0m.\n')
      exit()

# SP   
if __name__ == '__main__':
    main()