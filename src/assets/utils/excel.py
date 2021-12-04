from assets.utils.console import clearConsole, eraseLastLine, printListItem
import time, tkinter, sys
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import xlrd


def getExcelPath(fileName):
  root = tkinter.Tk()
  root.withdraw()
  root.wm_attributes('-topmost', 1)
  print('[ ] Select '+fileName+' File in dialog')
  time.sleep(.8)
  filePath = askopenfilename(filetypes = [( 'Excel File', '.xlsx .xls')])
  root.destroy()
  clearConsole()
  if(filePath == ''):
    sys.exit()
  print('[\033[92mx\033[0m] Select '+fileName+' File in dialog')
  time.sleep(2)
  return filePath

def getExcelSheet(filePath, fileName, sheetName, dataOnly = False):
  try:
    print(" |\n[ ] Loading data from "+"\033[92m.../"+filePath.split('/')[-2]+'/'+filePath.split('/')[-1]+"\033[0m")
    if dataOnly:
      wb = load_workbook(filePath, data_only=True)
    else:
      wb = load_workbook(filePath)
    eraseLastLine()
    print("[\033[92mx\033[0m] Loading data from "+"\033[92m.../"+filePath.split('/')[-2]+'/'+filePath.split('/')[-1]+"\033[0m\n")
    return wb[sheetName]
  except:
    clearConsole()
    printListItem('Something went wrong while loading the '+fileName+' File.', 'red')
    print('\nMake sure ...')
    printListItem('you \033[93mselect the right file.\033[0m', 'yellow')
    printListItem('you \033[93mclose the file\033[0m before loading.', 'yellow')
    printListItem('there is a \033[93m"'+sheetName+'"\033[0m Sheet.','yellow')
    print('')
    return None

def getProcessFileData(filePath, fileName):
  try:
    print(" |\n[ ] Loading data from "+"\033[92m.../"+filePath.split('/')[-2]+'/'+filePath.split('/')[-1]+"\033[0m")
    book = xlrd.open_workbook(filePath)
    eraseLastLine()
    print("[\033[92mx\033[0m] Loading data from "+"\033[92m.../"+filePath.split('/')[-2]+'/'+filePath.split('/')[-1]+"\033[0m\n")
    return book
  except:
    clearConsole()
    printListItem('Something went wrong while loading the '+fileName+' File.', 'red')
    print('\nMake sure ...')
    printListItem('you \033[93mclose the file\033[0m before running the script.', 'yellow')
    print('')
    return None