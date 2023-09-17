from openpyxl import Workbook
from assets.utils.programSteps import *

def main():

  desktop_path = initialActions()
  
# config data cycle
  while True:
    config_ws = getConfigSheet()
    config_data = getConfigData(config_ws)
    if config_data == 'load again': continue
    config_errors = getConfigErrors(config_data)
    if len(config_errors) > 0:
      if config_errors[0] == 'warnings':
        user_action = handleConfigWarnings(config_errors)
        if user_action == 'load again': continue
      else:
        user_action = handleConfigErrors()
        if user_action == 'load again': continue
    areaParameters       = config_data['areaParameters']
    typicalParameters    = config_data['typicalParameters']
    indexParameters      = config_data['indexParameters']
    processLibParameters = config_data['processLibParameters']
    break

  index_ws = getInstrumentIndexSheet()
  
# logic cycle
  while True:
    # get P&ID & typical functions
    # eventually all filtered typical lists with warnings
    while True:
      pid, allElements = getPidAndIndexElements(index_ws, indexParameters)
      if pid == 'load again': continue
      printPidFilterResult(allElements, pid, 'green')
      finalLists, allFilterWarnings = getFinalLists(allElements, typicalParameters, areaParameters, indexParameters)
      if finalLists == 'try again': continue
      else: break

    getMergeContinuationConfirmation()
    processLib_wb = getProcessLibWorkbook()
    output_wb = Workbook()
    
    # loop through entire final list of typical lists
    # merge with processLib sheets into output book
    for list in finalLists:
      signal_list, merge_function, typical_name = list
      
      sheet, rowCount = getProcessLibTypicalSheet(processLib_wb, typical_name)
      processLibSheetData = getProcessLibSheetData(sheet, rowCount)
      mergedOutput, controlOutput = callMergeFunction(merge_function, signal_list, processLibSheetData, processLibParameters[typical_name], typical_name)
      allFilterWarnings = removeNotUsedRangeWarnings(allFilterWarnings, mergedOutput, typical_name, processLibParameters[typical_name]['key'])
      
      if pid == 'all':
        # 0 is index of Full Tag in control sheet - filters normally return full tag at index 0
        differenceWarnings = returnMergeDifferences(mergedOutput, controlOutput, processLibParameters[typical_name]['key'], 0)
        allFilterWarnings = mergeCollectedWarnings(allFilterWarnings, differenceWarnings, typical_name)
      
      printListItem('Creating '+typical_name+' Sheet and importing Data ...', 'cyan')
      
      # create and populate control sheet
      controlSheet = output_wb.create_sheet(typical_name+'-Imported')
      controlSheet.append([' '])
      controlSheet.append(['This is just the control sheet, if you want to check the master imports for '+typical_name+' again.'])
      controlSheet.merge_cells('A1:J1')
      controlSheet.append(['The '+typical_name+'-Full Sheet has to be manually imported into the ProcessLib-File '+typical_name+' Sheet.'])
      controlSheet.merge_cells('A2:J2')
      controlSheet.append([' '])
      for row in controlOutput:
        controlSheet.append(row)
      
      # create and populate output sheet 
      outputSheet = output_wb.create_sheet(typical_name+'-Full')
      for row in mergedOutput:
        outputSheet.append(row)
    
    # save output file and display all warnings
    output_wb = removeDefaultSheet(output_wb)
    saveOutputFile(output_wb, desktop_path, pid)
    handleFilterWarnings(allFilterWarnings)
    
    # program end - user can restart
    confirmation = getUserConfirmation('Want to start again ?')
    clearConsole()
    if confirmation == True: continue
    else: break

# SP - NS Init   
if __name__ == '__main__':
  try:
    main()
  except KeyboardInterrupt:
    clearConsole()
    sys.exit()
  except Exception as issue:
    clearConsole()
    printListItem('Application crashed while facing the following issue:', 'red')
    print(issue)
    print('')